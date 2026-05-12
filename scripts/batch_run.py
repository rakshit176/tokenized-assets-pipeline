#!/usr/bin/env python3
"""Batch processor — run pipeline on multiple companies from CSV.

CSV format (no header, or with header):
    company_name,domain
    Securitize,securitize.io
    Ondo Finance,ondo.finance
    Centrifuge,centrifuge.io

Usage:
    python scripts/batch_run.py companies.csv [--workers N]
"""
import asyncio
import csv
import sys
import pathlib
import time
from typing import List, Tuple

sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))

from step_run import run_company


async def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/batch_run.py companies.csv [--workers N]")
        print("\nCSV format (no header):")
        print("  company_name,domain")
        print("  Securitize,securitize.io")
        print("  Ondo Finance,ondo.finance")
        print("\nOr with header:")
        print("  company_name,domain")
        print("  Securitize,securitize.io")
        sys.exit(1)

    csv_path = pathlib.Path(sys.argv[1])
    if not csv_path.exists():
        print(f"Error: {csv_path} does not exist")
        sys.exit(1)

    # Parse companies from CSV
    companies: List[Tuple[str, str]] = []
    with open(csv_path, newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) >= 2 and row[0] and row[1]:
                # Skip header row if present
                if row[0].lower() in ("company_name", "company") and row[1].lower() == "domain":
                    continue
                companies.append((row[0], row[1]))

    if not companies:
        print("Error: No valid company rows found in CSV")
        sys.exit(1)

    print(f"Found {len(companies)} companies to process")
    for name, domain in companies:
        print(f"  - {name} ({domain})")

    # Check for workers flag
    workers = 1
    if "--workers" in sys.argv:
        idx = sys.argv.index("--workers")
        if idx + 1 < len(sys.argv):
            workers = int(sys.argv[idx + 1])

    if workers > 1:
        # Parallel processing
        print(f"\nRunning {len(companies)} companies with {workers} parallel workers...")
        results = await asyncio.gather(*[run_company(name, domain) for name, domain in companies])
    else:
        # Sequential processing
        print(f"\nRunning {len(companies)} companies sequentially...")
        results = []
        for name, domain in companies:
            result = await run_company(name, domain)
            results.append(result)

    # Print summary
    print(f"\n{'='*60}")
    print(f"  BATCH COMPLETE: {len(companies)} companies")
    print(f"{'='*60}")
    for r in results:
        print(f"  {r['company']:20s} | Fill={r['fill_rate']:5.1f}% Real={r['real_fill']:5.1f}% "
              f"Cost=${r['total_cost_usd']:.4f} Time={r['timing']['total']:.1f}s")

    # Write comparison Excel
    if len(results) > 1:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Batch Results"

        hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF")

        headers = ["Company", "Domain", "Fill Rate", "Real Fill", "High Fill",
                   "Tables", "Total Tokens", "Cost (USD)", "Search (s)",
                   "Scrape (s)", "LLM (s)", "Gap Fill (s)", "Total (s)"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        for r in results:
            ws.append([
                r["company"], r["domain"],
                f"{r['fill_rate']}%", f"{r['real_fill']}%", f"{r['high_fill']}%",
                r["tables_with_data"], r["total_tokens"],
                f"${r['total_cost_usd']:.4f}",
                r["timing"]["search"], r["timing"]["scrape"],
                r["timing"]["llm"], r["timing"]["gap_fill"],
                r["timing"]["total"],
            ])

        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
            ws.column_dimensions[col].width = 16

        output_path = pathlib.Path("output/batch_results.xlsx")
        output_path.parent.mkdir(exist_ok=True)
        wb.save(str(output_path))
        print(f"\n  Batch results: {output_path}")


if __name__ == "__main__":
    asyncio.run(main())
