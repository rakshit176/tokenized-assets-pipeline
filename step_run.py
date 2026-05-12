#!/usr/bin/env python3
"""Run the full extraction pipeline on one or more companies.

Usage:
    python step_run.py                                    # default: Securitize
    python step_run.py Securitize securitize.io
    python step_run.py "Ondo Finance" ondo.finance Centrifuge centrifuge.io
"""
import asyncio
import json
import os
import sys
import time
import pathlib
from typing import Any

from dotenv import load_dotenv
load_dotenv(pathlib.Path(__file__).parent / ".env")

CHILD_TABLES = [
    'products', 'asset_classes', 'features', 'standards', 'integrations',
    'funding_rounds', 'governance_models', 'compliance_certifications',
    'sla_commitments', 'stability_entries', 'tech_stack', 'api_capabilities',
    'partnerships', 'exchange_listings', 'regulatory_licenses',
    'platform_metrics', 'deal_case_studies',
]


def _flatten_schema(cd) -> list[dict]:
    """Flatten CompanyData into rows: table, field, value, source_url, confidence."""
    rows = []
    from src.schema.models import CitedValue

    def _extract(table_name: str, model, row_idx: int | None = None):
        for fn in model.__class__.model_fields:
            val = getattr(model, fn)
            if isinstance(val, CitedValue):
                rows.append({
                    "table": table_name,
                    "row": row_idx if row_idx is not None else "",
                    "field": fn,
                    "value": str(val.value) if val.value is not None else "",
                    "source_url": val.source_url or "",
                    "confidence": round(val.confidence, 2),
                })

    _extract("companies", cd.profile)
    for t_name in CHILD_TABLES:
        items = getattr(cd, t_name, [])
        for i, item in enumerate(items):
            _extract(t_name, item, row_idx=i)

    return rows


def _write_excel(company_name: str, domain: str, cd, call_logs: list[dict],
                 metrics: dict, output_dir: pathlib.Path):
    """Write a multi-sheet Excel workbook with schema, LLM costs, and evaluation."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # --- Sheet 1: Schema Data ---
    ws1 = wb.active
    ws1.title = "Schema Data"
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    schema_rows = _flatten_schema(cd)
    headers1 = ["Table", "Row", "Field", "Value", "Source URL", "Confidence"]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in schema_rows:
        ws1.append([r["table"], r["row"], r["field"], r["value"], r["source_url"], r["confidence"]])

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 6
    ws1.column_dimensions["C"].width = 28
    ws1.column_dimensions["D"].width = 45
    ws1.column_dimensions["E"].width = 50
    ws1.column_dimensions["F"].width = 12

    # --- Sheet 2: LLM Cost ---
    ws2 = wb.create_sheet("LLM Cost")
    headers2 = ["Batch", "Provider", "Model", "Prompt Tokens", "Completion Tokens",
                "Total Tokens", "Cost (USD)", "Latency (ms)", "User Prompt Chars",
                "Response Chars", "Success", "Error"]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = header_font_white
        cell.fill = headerFill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for log in call_logs:
        ws2.append([
            log.get("batch_name", ""),
            log.get("provider", ""),
            log.get("model", ""),
            log.get("prompt_tokens", 0),
            log.get("completion_tokens", 0),
            log.get("total_tokens", 0),
            log.get("cost_usd", 0),
            log.get("latency_ms", 0),
            log.get("user_prompt_chars", 0),
            log.get("response_chars", 0),
            log.get("success", True),
            log.get("error_message", ""),
        ])

    # Summary row
    total_pt = sum(l.get("prompt_tokens", 0) for l in call_logs)
    total_ct = sum(l.get("completion_tokens", 0) for l in call_logs)
    total_tt = sum(l.get("total_tokens", 0) for l in call_logs)
    total_cost = sum(l.get("cost_usd", 0) for l in call_logs)
    total_lat = sum(l.get("latency_ms", 0) for l in call_logs)
    ws2.append([])
    ws2.append(["TOTAL", "", "", total_pt, total_ct, total_tt,
                round(total_cost, 6), total_lat, "", "",
                f"{len(call_logs)} calls", ""])

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 22
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 18
    ws2.column_dimensions["F"].width = 14
    ws2.column_dimensions["G"].width = 14
    ws2.column_dimensions["H"].width = 14
    ws2.column_dimensions["I"].width = 18
    ws2.column_dimensions["J"].width = 16
    ws2.column_dimensions["K"].width = 10
    ws2.column_dimensions["L"].width = 20

    # --- Sheet 3: Evaluation ---
    ws3 = wb.create_sheet("Evaluation")
    eval_data = [
        ["Metric", "Value"],
        ["Company", company_name],
        ["Domain", domain],
        ["LLM Provider", metrics.get("llm_provider", "")],
        ["LLM Model", metrics.get("llm_model", "")],
        ["Overall Fill Rate", f"{metrics.get('fill_rate', 0)*100:.1f}%"],
        ["Real Fill Rate (conf>=0.4)", f"{metrics.get('real_fill', 0)*100:.1f}%"],
        ["High Fill Rate (conf>=0.7)", f"{metrics.get('high_fill', 0)*100:.1f}%"],
        ["Tables with Data", metrics.get("tables_with_data", "")],
        ["Total Fields", metrics.get("total_fields", "")],
        ["Real Filled Fields", metrics.get("real_filled_fields", "")],
        ["Search Results", metrics.get("search_results", 0)],
        ["Pages Scraped", metrics.get("pages_scraped", 0)],
        ["LLM Calls", len(call_logs)],
        ["Total Tokens", total_tt],
        ["Total LLM Cost (USD)", f"${total_cost:.6f}"],
        ["Total Time (seconds)", f"{metrics.get('duration_secs', 0):.1f}"],
        ["Search Time (seconds)", f"{metrics.get('search_time', 0):.1f}"],
        ["Scrape Time (seconds)", f"{metrics.get('scrape_time', 0):.1f}"],
        ["LLM Time (seconds)", f"{metrics.get('llm_time', 0):.1f}"],
        ["Gap Fill Time (seconds)", f"{metrics.get('gap_time', 0):.1f}"],
        [],
        ["Per-Table Breakdown", "Fields", "Filled (conf>=0.4)", "Fill %"],
    ]
    for row in eval_data:
        ws3.append(row)

    for cell in ws3[1]:
        cell.font = header_font_white
        cell.fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Per-table breakdown
    from src.schema.models import CitedValue
    table_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    def _table_stats(table_name, model_or_list):
        if isinstance(model_or_list, list):
            total_f = 0
            filled_f = 0
            for item in model_or_list:
                for fn in item.__class__.model_fields:
                    val = getattr(item, fn)
                    if isinstance(val, CitedValue):
                        total_f += 1
                        if val.is_filled() and val.confidence >= 0.4:
                            filled_f += 1
            return total_f, filled_f
        else:
            total_f = 0
            filled_f = 0
            for fn in model_or_list.__class__.model_fields:
                val = getattr(model_or_list, fn)
                if isinstance(val, CitedValue):
                    total_f += 1
                    if val.is_filled() and val.confidence >= 0.4:
                        filled_f += 1
            return total_f, filled_f

    for t_name in ["companies"] + CHILD_TABLES:
        if t_name == "companies":
            tf, ff = _table_stats("companies", cd.profile)
        else:
            tf, ff = _table_stats(t_name, getattr(cd, t_name, []))
        pct = f"{(ff/tf*100):.1f}%" if tf else "0%"
        ws3.append([t_name, tf, ff, pct])

    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 25
    ws3.column_dimensions["C"].width = 22
    ws3.column_dimensions["D"].width = 12

    # Save
    safe_name = domain.replace(".", "_")
    path = output_dir / f"{safe_name}.xlsx"
    wb.save(str(path))
    print(f"  Excel: {path}", flush=True)


async def run_company(company_name: str, domain: str):
    """Run the full pipeline for a single company."""
    from src.search.client import SearXNGClient
    from src.scrape.scraper import AsyncScraper
    from src.extractor.llm import LLMExtractor
    from src.orchestrator import PipelineRunner
    from src.schema.models import CompanyData, CitedValue
    from src.cache import get_cache

    t0 = time.time()
    print(f"\n{'='*60}\n  {company_name} ({domain})\n{'='*60}", flush=True)

    # Step 1: Search
    print("  [1/5] Search...", end=" ", flush=True)
    t1 = time.time()
    s = SearXNGClient()
    sr = await s.deep_search(company_name, domain)
    search_time = time.time() - t1
    print(f"{sum(len(v) for v in sr.values())} results ({search_time:.1f}s)", flush=True)

    # Step 2: Scrape
    print("  [2/5] Scrape...", end=" ", flush=True)
    t2 = time.time()
    scraper = AsyncScraper(timeout=10, max_concurrent=8, use_playwright=True)
    cp = await scraper.get_company_pages(domain)
    sp = await scraper.scrape_search_urls(sr, max_per_query=2, max_total=10, exclude_domain=domain)
    subdomain_urls = []
    for q, results in sr.items():
        for r in results:
            url = r.get("url", "") if isinstance(r, dict) else getattr(r, "url", "")
            if url and domain in url and url not in cp and url not in sp:
                subdomain_urls.append(url)
    if subdomain_urls:
        sdp = await scraper.scrape_urls(subdomain_urls[:8])
        sp.update(sdp)
    all_pages = {**cp, **sp}
    scrape_time = time.time() - t2
    print(f"{len(all_pages)} pages ({scrape_time:.1f}s)", flush=True)

    # Step 3: LLM extraction
    print("  [3/5] LLM extraction (6 batches)...", end=" ", flush=True)
    t3 = time.time()
    ext = LLMExtractor()
    ex = await ext.extract_all(company_name, domain, sr, all_pages)
    llm_time = time.time() - t3
    print(f"done ({llm_time:.1f}s)", flush=True)

    cd = CompanyData()
    PipelineRunner._merge_data(cd, PipelineRunner._structure_output(ex, company_name, domain))
    cd.apply_defaults(company_name, domain)

    # Step 4: Knowledge gap fill
    print("  [4/5] Knowledge gap fill...", end=" ", flush=True)
    t4 = time.time()
    missing = cd.missing_fields(0.4)
    if missing:
        kg = await ext.extract_knowledge_gaps(company_name, domain, missing)
        if kg:
            kg_data = PipelineRunner._structure_output(kg, company_name, domain)
            PipelineRunner._merge_data(cd, kg_data)
        cd.apply_defaults(company_name, domain)
    gap_time = time.time() - t4
    print(f"{len(missing)} gaps filled ({gap_time:.1f}s)", flush=True)

    # Calculate metrics
    rf = cd.real_fill_score(0.4)
    hf = cd.real_fill_score(0.7)
    ff = cd.confidence_score()
    elapsed = time.time() - t0

    # Count fields
    total_fields = 0
    real_filled = 0
    for fn in cd.profile.__class__.model_fields:
        val = getattr(cd.profile, fn)
        if isinstance(val, CitedValue):
            total_fields += 1
            if val.is_filled() and val.confidence >= 0.4:
                real_filled += 1
    for t_name in CHILD_TABLES:
        for item in getattr(cd, t_name, []):
            for fn in item.__class__.model_fields:
                val = getattr(item, fn)
                if isinstance(val, CitedValue):
                    total_fields += 1
                    if val.is_filled() and val.confidence >= 0.4:
                        real_filled += 1

    tables_with_data = sum(1 for t in CHILD_TABLES if len(getattr(cd, t, [])) > 0)

    call_logs = getattr(ext, 'call_logs', [])
    # Calculate cost per log
    from src.database.saver import DatabaseSaver
    for log in call_logs:
        log["cost_usd"] = DatabaseSaver._calc_cost(
            log.get("model", ""), log.get("prompt_tokens", 0), log.get("completion_tokens", 0)
        )

    total_tokens = sum(l.get("total_tokens", 0) for l in call_logs)
    total_cost = sum(l.get("cost_usd", 0) for l in call_logs)

    metrics = {
        "llm_provider": os.getenv("LLM_PROVIDER", "zai"),
        "llm_model": ext.provider.config.model,
        "fill_rate": ff,
        "real_fill": rf,
        "high_fill": hf,
        "tables_with_data": f"{tables_with_data}/17",
        "total_fields": total_fields,
        "real_filled_fields": real_filled,
        "search_results": sum(len(v) for v in sr.values()),
        "pages_scraped": len(all_pages),
        "duration_secs": elapsed,
        "search_time": search_time,
        "scrape_time": scrape_time,
        "llm_time": llm_time,
        "gap_time": gap_time,
    }

    print(f"  [5/5] Saving...", flush=True)

    # JSON output
    output_dir = pathlib.Path("output")
    output_dir.mkdir(exist_ok=True)
    safe_name = domain.replace(".", "_")
    result = {
        "company": company_name,
        "domain": domain,
        "fill_rate": round(ff * 100, 1),
        "real_fill": round(rf * 100, 1),
        "high_fill": round(hf * 100, 1),
        "tables_with_data": f"{tables_with_data}/17",
        "total_fields": total_fields,
        "real_filled_fields": real_filled,
        "total_tokens": total_tokens,
        "total_cost_usd": round(total_cost, 6),
        "timing": {
            "total": round(elapsed, 1),
            "search": round(search_time, 1),
            "scrape": round(scrape_time, 1),
            "llm": round(llm_time, 1),
            "gap_fill": round(gap_time, 1),
        },
        "data": cd.model_dump(),
    }
    json_path = output_dir / f"{safe_name}.json"
    json_path.write_text(json.dumps(result, default=str, indent=2))

    # Excel output
    _write_excel(company_name, domain, cd, call_logs, metrics, output_dir)

    # Save to PostgreSQL
    try:
        from src.database.saver import DatabaseSaver
        db = DatabaseSaver()
        await db.save(
            company_name=company_name, domain=domain, data=cd,
            llm_provider=metrics["llm_provider"],
            llm_model=metrics["llm_model"],
            fill_rate=ff, real_fill=rf, high_fill=hf,
            pages_scraped=len(all_pages),
            search_results=metrics["search_results"],
            duration_secs=elapsed,
            llm_call_logs=call_logs,
        )
        await db.close()
        print(f"    DB: saved", flush=True)
    except Exception as e:
        print(f"    DB: failed ({e})", flush=True)

    await scraper.close()

    print(f"\n  RESULT: Fill={ff*100:.1f}% Real={rf*100:.1f}% High={hf*100:.1f}% "
          f"Tables={tables_with_data}/17 Tokens={total_tokens:,} "
          f"Cost=${total_cost:.4f} Time={elapsed:.1f}s\n", flush=True)

    return result


async def main():
    args = sys.argv[1:]

    # Default company
    if not args:
        companies = [("Securitize", "securitize.io")]
    else:
        companies = []
        i = 0
        while i < len(args):
            name = args[i]
            domain = args[i + 1] if i + 1 < len(args) else None
            if not domain:
                print(f"Error: provide both name and domain for '{name}'")
                print("Usage: python step_run.py [Company1 domain1.com] [Company2 domain2.com] ...")
                sys.exit(1)
            companies.append((name, domain))
            i += 2

    results = []
    for name, domain in companies:
        r = await run_company(name, domain)
        results.append(r)

    # Summary
    if len(results) > 1:
        print(f"\n{'='*60}")
        print(f"  SUMMARY: {len(results)} companies")
        print(f"{'='*60}")
        for r in results:
            print(f"  {r['company']:20s} | Fill={r['fill_rate']:5.1f}% Real={r['real_fill']:5.1f}% "
                  f"Cost=${r['total_cost_usd']:.4f} Time={r['timing']['total']:.1f}s", flush=True)

    # Comparison Excel
    if len(results) > 1:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparison"

        hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF")

        headers = ["Company", "Domain", "Fill Rate", "Real Fill", "High Fill",
                   "Tables", "Total Tokens", "Cost (USD)", "Search (s)",
                   "Scrape (s)", "LLM (s)", "Gap Fill (s)", "Total (s)"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        for r in results:
            ws.append([
                r["company"], r["domain"],
                f"{r['fill_rate']}%", f"{r['real_fill']}%", f"{r['high_fill']}%",
                r["tables_with_data"], r["total_tokens"],
                f"${r['total_cost_usd']:.4f}",
                r["timing"]["search"], r["timing"]["scrape"],
                r["timing"]["llm"], r["timing"]["gap_fill"],
                r["timing"]["total"],
            ])

        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]:
            ws.column_dimensions[col].width = 16

        wb.save("output/comparison.xlsx")
        print(f"\n  Comparison: output/comparison.xlsx", flush=True)

    # Cache stats
    try:
        cache = get_cache()
        stats = cache.stats()
        if stats["hits"] + stats["misses"] > 0:
            print(f"\n  Cache: {stats['hits']} hits, {stats['misses']} misses, "
                  f"{stats['hit_rate']} hit rate, {stats['cache_files']} files", flush=True)
    except Exception:
        pass


if __name__ == "__main__":
    asyncio.run(main())
